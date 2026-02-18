"""ETL script for Motilal Oswal Mutual Fund portfolio data.

Extracts equity holdings from monthly portfolio Excel files into structured JSON.

Column mapping (0-indexed):
- 0: Sr. No.
- 1: Name of Instrument
- 2: (empty/spacing)
- 3: ISIN
- 4: Rating / Industry
- 5: Quantity
- 6: Market value (Rs. In lakhs)
- 7: % to Net Assets
"""

import argparse
import json
import logging
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

# Column indices (0-indexed)
COL_SR_NO = 0
COL_NAME = 1
COL_ISIN = 3
COL_INDUSTRY = 4
COL_QUANTITY = 5
COL_MARKET_VALUE = 6
COL_PCT_NAV = 7

# Fallback scheme name mapping for files without an Index sheet.
# Built from files that do have the Index sheet (Jan/Feb/Nov 2025, Jan 2026).
FALLBACK_SCHEME_MAP = {
    "YO01": "Motilal Oswal Nifty 50 ETF (Formerly known as Motilal Oswal M50 ETF)",
    "YO02": "Motilal Oswal Nifty Midcap 100 ETF (Formerly known as Motilal Oswal Midcap 100 ETF)",
    "YO05": "Motilal Oswal Focused Fund (Formerly known as Motilal Oswal Focused 25 Fund)",
    "YO07": "Motilal Oswal Midcap Fund (Formerly known as Motilal Oswal Midcap 30 Fund)",
    "YO08": "Motilal Oswal Flexi Cap Fund",
    "YO09": "Motilal Oswal ELSS Tax Saver Fund (Formerly Known as Motilal Oswal Long Term Equity Fund)",
    "YO10": "Motilal Oswal Balanced Advantage Fund (Formerly known as Motilal Oswal Dynamic Fund)",
    "YO16": "Motilal Oswal Nifty Midcap 150 Index Fund",
    "YO17": "Motilal Oswal Nifty Smallcap 250 Index Fund",
    "YO18": "Motilal Oswal Nifty 500 Index Fund (Formerly known as Motilal Oswal Nifty 500 Fund)",
    "YO19": "Motilal Oswal Nifty Bank Index Fund",
    "YO20": "Motilal Oswal Large and Midcap Fund",
    "YO21": "Motilal Oswal Nifty 50 Index Fund",
    "YO22": "Motilal Oswal Nifty Next 50 Index Fund",
    "YO24": "Motilal Oswal Multi Asset Fund",
    "YO31": "Motilal Oswal Nifty 200 Momentum 30 ETF",
    "YO32": "Motilal Oswal Nifty 200 Momentum 30 Index Fund",
    "YO33": "Motilal OswalBSE Low Volatility ETF",
    "YO34": "Motilal OswalBSE Low Volatility Index Fund",
    "YO35": "Motilal OswalBSE Financials ex Bank 30 Index Fund",
    "YO36": "Motilal OswalBSE Healthcare ETF",
    "YO37": "Motilal OswalBSE Enhanced Value ETF",
    "YO38": "Motilal OswalBSE Enhanced Value Index Fund",
    "YO39": "Motilal OswalBSE Quality ETF",
    "YO40": "Motilal OswalBSE Quality Index Fund",
    "YO43": "Motilal Oswal Nifty Microcap 250 Index Fund",
    "YO45": "Motilal Oswal Nifty 500 ETF",
    "YO46": "Motilal Oswal Small Cap Fund",
    "YO47": "Motilal Oswal Large Cap Fund",
    "YO48": "Motilal Oswal Nifty Realty ETF",
    "YO49": "Motilal Oswal Nifty Smallcap 250 ETF",
    "YO50": "Motilal Oswal Quant Fund",
    "YO51": "Motilal Oswal Multicap Fund",
    "YO52": "Motilal Oswal Nifty India Defence Index Fund",
    "YO53": "Motilal Oswal Manufacturing Fund",
    "YO54": "Motilal Oswal Business Cycle Fund",
    "YO55": "Motilal Oswal Nifty India Defence ETF",
    "YO56": "Motilal Oswal Nifty 500 Momentum 50 Index Fund",
    "YO57": "Motilal Oswal Nifty 500 Momentum 50 ETF",
    "YO58": "Motilal Oswal Digital India Fund",
    "YO59": "Motilal Oswal Nifty MidSmall Fin Servs Index Fund",
    "YO60": "MO Nifty MidSmall India Consumption Index Fund",
    "YO61": "Motilal Oswal Nifty MidSmall Healthcare Index Fund",
    "YO62": "MO Nifty MidSmall IT and Telecom Index Fund",
    "YO63": "MO Nifty Capital Market Index Fund",
    "YO64": "Motilal Oswal Arbitrage Fund",
    "YO65": "Motilal Oswal Innovation Opportunities Fund",
    "YO66": "Motilal Oswal Active Momentum Fund",
    "YO67": "Motilal Oswal Nifty Capital Market ETF",
    "YO68": "Motilal Oswal Infrastructure Fund",
    "YO69": "Motilal Oswal Nifty 50 Equal Weight ETF",
    "YO70": "Motilal Oswal Nifty Next 50 ETF",
    "YO71": "Motilal Oswal BSE India Infrastructure ETF",
    "YO72": "Motilal Oswal Services Fund",
    "YO73": "Motilal Oswal Nifty India Manufacturing ETF",
    "YO74": "Motilal Oswal Nifty PSE ETF",
    "YO75": "Motilal Oswal Nifty India Tourism ETF",
    "YO76": "Motilal Oswal Nifty Midcap 150 Momentum 50 ETF",
    "YO77": "Motilal Oswal Nifty Alpha 50 ETF",
    "YO78": "Motilal Oswal BSE 1000 Index Fund",
    "YO80": "Motilal Oswal Special Opportunities Fund",
    "YO82": "Motilal Oswal Consumption Fund",
    "YO83": "Motilal Oswal Nifty 100 ETF",
    "YO84": "Motilal Oswal Nifty Energy ETF",
    "YO85": "Motilal Oswal BSE Select IPO ETF",
    "YO86": "Motilal Oswal Nifty Services Sector ETF",
    "YO87": "Motilal Oswal Nifty MNC ETF",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("motilal_etl")


def safe_float(val: Any) -> float | None:
    """Convert value to float, return None if invalid."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "NIL", "N/A", "#", "-", "Total", "Sub Total"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def clean_str(val: Any) -> str | None:
    """Clean string value, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in ("", "NIL", "N/A", "#", "-") else None


def parse_date(val: Any) -> str | None:
    """Parse date to YYYY-MM-DD format."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    
    # Clean up common issues
    s = s.replace(",", ", ")
    s = re.sub(r'\s+', ' ', s)
    
    # Try common formats
    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%B %d, %Y",  # January 31, 2026
        "%B %d,%Y",   # January 31,2026
        "%d-%b-%Y",   # 31-Jan-2026
        "%d %B %Y",   # 31 January 2026
    )
    
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    # Try with title case
    s_title = s.title()
    for fmt in formats:
        try:
            return datetime.strptime(s_title, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    return s


def parse_index_sheet(wb: openpyxl.Workbook) -> dict[str, str]:
    """Parse Index sheet to map sheet codes to scheme names.
    
    Index sheet structure:
    - Row 4: Headers ("Sr No." in Col C, "Fund Name" in Col D, "Fund Code" in Col E)
    - Row 5+: Data rows
    - Column D (index 3): Scheme name
    - Column E (index 4): Sheet code (e.g., "YO01", "YO02")
    
    Returns:
        dict mapping sheet code to scheme name
    """
    if "Index" not in wb.sheetnames:
        log.warning("No 'Index' sheet found")
        return {}
    
    ws = wb["Index"]
    rows = list(ws.iter_rows(values_only=True))
    
    scheme_map = {}
    
    # Skip header rows (row 4 is headers with "Sr No.", "Fund Name", "Fund Code")
    # Data starts from row 5 (index 4)
    for i, r in enumerate(rows[4:], start=5):
        # Column D (index 3) has scheme name
        scheme_name = clean_str(r[3]) if len(r) > 3 else None
        
        # Column E (index 4) has sheet code
        sheet_code = clean_str(r[4]) if len(r) > 4 else None
        
        if sheet_code and scheme_name:
            scheme_map[sheet_code] = scheme_name
            log.debug(f"  {sheet_code} -> {scheme_name}")
    
    log.info(f"Parsed Index sheet: {len(scheme_map)} schemes")
    return scheme_map


def extract_equity_holdings(ws_rows: list[tuple], scheme_info: dict) -> dict | None:
    """Extract listed equity holdings from a scheme sheet.
    
    Args:
        ws_rows: All rows from the worksheet
        scheme_info: Dict with scheme_code and scheme_name
        
    Returns:
        Dict with scheme info and holdings list, or None if no equity data
    """
    holdings = []
    equity_section_row = None
    listed_section_row = None
    
    # Find equity section markers (check column 0 and 1 like other AMCs)
    for i, r in enumerate(ws_rows):
        cell_col0 = clean_str(r[0]) if r else None
        cell_col1 = clean_str(r[1]) if len(r) > 1 else None
        
        # Check for equity markers
        if cell_col0 and "EQUITY" in cell_col0.upper() and "RELATED" in cell_col0.upper():
            equity_section_row = i
            log.debug(f"  Found equity section at row {i+1} (col 0)")
        elif cell_col1 and "EQUITY" in cell_col1.upper() and "RELATED" in cell_col1.upper():
            equity_section_row = i
            log.debug(f"  Found equity section at row {i+1} (col 1)")
        
        # Check for "Listed" subsection
        if equity_section_row is not None:
            if cell_col0 and "LISTED" in cell_col0.upper():
                listed_section_row = i
                log.debug(f"  Found listed section at row {i+1}")
                break
            elif cell_col1 and "LISTED" in cell_col1.upper():
                listed_section_row = i
                log.debug(f"  Found listed section at row {i+1}")
                break
    
    if listed_section_row is None:
        log.debug(f"  No listed equity section found in {scheme_info['scheme_code']}")
        return None
    
    # Extract holdings starting after listed section marker
    for i in range(listed_section_row + 1, len(ws_rows)):
        r = ws_rows[i]
        
        # Extract data first
        name = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        isin = clean_str(r[COL_ISIN]) if len(r) > COL_ISIN else None
        industry = clean_str(r[COL_INDUSTRY]) if len(r) > COL_INDUSTRY else None
        quantity = safe_float(r[COL_QUANTITY]) if len(r) > COL_QUANTITY else None
        market_value = safe_float(r[COL_MARKET_VALUE]) if len(r) > COL_MARKET_VALUE else None
        pct_nav = safe_float(r[COL_PCT_NAV]) if len(r) > COL_PCT_NAV else None
        
        # Check for end marker in Name column ONLY (where "Total" actually appears)
        # This prevents premature stopping due to empty cells in other columns
        if name:
            name_lower = name.lower().strip()
            # End conditions: "Total", "Sub Total", or section (b)
            if name_lower == "total" or name_lower == "sub total" or name_lower.startswith("(b)"):
                break
        
        # Valid holding must have ISIN and name
        if not isin or not name:
            continue
        
        # ISIN validation
        if len(isin) != 12 or not isin.startswith("INE"):
            log.debug(f"  Skipping invalid ISIN: {isin}")
            continue
        
        holdings.append({
            "isin": isin,
            "security_name": name,
            "industry": industry,
            "quantity": quantity,
            "market_value_lakhs": market_value,
            "pct_to_nav": pct_nav,
        })
    
    if not holdings:
        return None
    
    # Aggregate duplicate ISINs
    aggregated_holdings = {}
    for h in holdings:
        isin = h["isin"]
        if isin in aggregated_holdings:
            existing = aggregated_holdings[isin]
            if h["quantity"] is not None:
                existing["quantity"] = (existing["quantity"] or 0) + h["quantity"]
            if h["market_value_lakhs"] is not None:
                existing["market_value_lakhs"] = (existing["market_value_lakhs"] or 0) + h["market_value_lakhs"]
            if h["pct_to_nav"] is not None:
                existing["pct_to_nav"] = (existing["pct_to_nav"] or 0) + h["pct_to_nav"]
        else:
            aggregated_holdings[isin] = h
    
    final_holdings = list(aggregated_holdings.values())
    
    return {
        "scheme_short_code": scheme_info["scheme_code"],
        "scheme_name": scheme_info["scheme_name"],
        "holdings": final_holdings,
    }


def run_etl(excel_path: str, date_override: str | None = None) -> dict:
    """Run the ETL process on a Motilal Oswal portfolio Excel file."""
    log.info(f"Processing: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # Parse index sheet
    scheme_map = parse_index_sheet(wb)
    
    # Extract report date
    report_date = None
    
    if date_override:
        report_date = date_override
        log.info(f"Using provided report date: {report_date}")
    elif len(wb.sheetnames) > 1:
        # Try to extract from first sheet
        first_sheet = wb.sheetnames[1]
        ws = wb[first_sheet]
        rows = list(ws.iter_rows(max_row=10, values_only=True))
        for r in rows:
            for cell in r[:5]:
                if cell:
                    parsed = parse_date(cell)
                    if parsed and parsed != str(cell):
                        report_date = parsed
                        break
            if report_date:
                break
    
    if not report_date:
        log.warning("Could not extract report date, using current date")
        report_date = datetime.now().strftime("%Y-%m-%d")
    
    log.info(f"Report date: {report_date}")
    
    # Process each scheme sheet
    all_scheme_data = []
    isin_counter = Counter()
    total_holdings = 0
    errors = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue
        
        # Get scheme name: Index sheet > fallback map > row 1 col B
        scheme_name = scheme_map.get(sheet_name)
        
        if not scheme_name:
            scheme_name = FALLBACK_SCHEME_MAP.get(sheet_name)
        
        try:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            # If still no name, try row 1, column B (index 1)
            if not scheme_name and rows:
                row1_colb = clean_str(rows[0][1]) if len(rows[0]) > 1 else None
                if row1_colb and "motilal" in row1_colb.lower():
                    scheme_name = row1_colb
                    log.info(f"  Got scheme name from row 1 col B: {scheme_name}")
            
            if not scheme_name:
                scheme_name = f"Unknown Scheme ({sheet_name})"
            
            scheme_info = {
                "scheme_code": sheet_name,
                "scheme_name": scheme_name,
            }
            
            result = extract_equity_holdings(rows, scheme_info)
            
            if result:
                all_scheme_data.append(result)
                holdings_count = len(result["holdings"])
                total_holdings += holdings_count
                
                for h in result["holdings"]:
                    isin_counter[h["isin"]] += 1
                
                log.info(f"  ✓ {sheet_name}: {holdings_count} holdings")
        except Exception as e:
            log.error(f"  ✗ {sheet_name}: {e}")
            errors += 1
    
    wb.close()
    
    # Build output structure
    output = {
        "metadata": {
            "source_file": Path(excel_path).name,
            "report_date": report_date,
            "extraction_date": datetime.now().strftime("%Y-%m-%d"),
            "total_schemes": len(wb.sheetnames) - 1,  # Exclude Index
            "schemes_with_equity": len(all_scheme_data),
            "schemes_skipped": len(wb.sheetnames) - 1 - len(all_scheme_data),
            "errors": errors,
            "total_unique_securities": len(isin_counter),
            "total_holdings_records": total_holdings,
        },
        "amc_master": {
            "amc_name": "Motilal Oswal Mutual Fund",
            "short_code": "MOTILAL",
        },
        "fund_master": [],
        "security_master": [],
        "portfolio_holdings": [],
    }
    
    # Build fund_master and security_master
    for scheme in all_scheme_data:
        output["fund_master"].append({
            "scheme_short_code": scheme["scheme_short_code"],
            "scheme_name": scheme["scheme_name"],
            "scheme_code": scheme["scheme_short_code"],
            "holdings_count": len(scheme["holdings"]),
        })
        
        for holding in scheme["holdings"]:
            output["portfolio_holdings"].append({
                "scheme_short_code": scheme["scheme_short_code"],
                "isin": holding["isin"],
                "quantity": holding["quantity"],
                "market_value_lakhs": holding["market_value_lakhs"],
                "pct_to_aum": holding["pct_to_nav"],
                "industry": holding["industry"],
            })
    
    # Build unique security master
    seen_isins = set()
    for holding in output["portfolio_holdings"]:
        isin = holding["isin"]
        if isin not in seen_isins:
            # Find first occurrence to get name and industry
            for scheme in all_scheme_data:
                for h in scheme["holdings"]:
                    if h["isin"] == isin:
                        output["security_master"].append({
                            "isin": isin,
                            "security_name": h["security_name"],
                            "current_industry": h["industry"],
                            "current_sector": None,
                        })
                        seen_isins.add(isin)
                        break
                if isin in seen_isins:
                    break
    
    log.info(f"Extraction complete: {len(all_scheme_data)} schemes, {total_holdings} holdings")
    return output


def main():
    parser = argparse.ArgumentParser(
        description="Extract equity holdings from Motilal Oswal Mutual Fund portfolio Excel"
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--date", "-d", help="Override report date (YYYY-MM-DD)")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Run ETL
    data = run_etl(args.excel_file, date_override=args.date)
    
    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        excel_name = Path(args.excel_file).stem
        output_path = Path("data/processed/motilal") / f"motilal_equity_holdings_{excel_name.lower()}.json"
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Write JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    log.info(f"Output written to: {output_path}")
    log.info(f"  Schemes: {data['metadata']['schemes_with_equity']}")
    log.info(f"  Unique securities: {data['metadata']['total_unique_securities']}")
    log.info(f"  Total holdings: {data['metadata']['total_holdings_records']}")


if __name__ == "__main__":
    main()
