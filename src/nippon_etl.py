"""
ETL script for Nippon India Mutual Fund monthly portfolio Excel files.
Extracts Listed Equity holdings and outputs JSON matching database schema.
"""

import argparse
import json
import logging
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

# Column indices for Nippon India (0-indexed)
COL_INTERNAL_CODE = 0   # Internal security code (optional)
COL_ISIN = 1            # ISIN
COL_NAME = 2            # Security Name
COL_INDUSTRY = 3        # Industry/Rating
COL_QUANTITY = 4        # Quantity
COL_MARKET_VALUE = 5    # Market/Fair Value (Rs. in Lacs)
COL_PCT_NAV = 6         # % to NAV
COL_YIELD = 7           # YIELD (optional)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("nippon_etl")


def safe_float(val: Any) -> float | None:
    """Convert value to float, return None if invalid."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "NIL", "N/A", "#", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def safe_int(val: Any) -> int | None:
    """Convert value to int, return None if invalid."""
    f = safe_float(val)
    return int(f) if f is not None else None


def clean_str(val: Any) -> str | None:
    """Clean string value, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in ("", "NIL", "N/A", "#", "-") else None


def parse_date(val: Any) -> str | None:
    """Parse date to YYYY-MM-DD format."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%B %d,%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def parse_index_sheet(wb: openpyxl.Workbook) -> list[dict]:
    """
    Parse Index sheet to get scheme metadata.
    Returns list of dicts with scheme_short_code and scheme_name.
    """
    # Handle both 'Index' and 'INDEX'
    index_sheet_name = None
    for name in ["Index", "INDEX"]:
        if name in wb.sheetnames:
            index_sheet_name = name
            break
    
    if not index_sheet_name:
        raise ValueError("Index sheet not found (tried 'Index' and 'INDEX')")
    
    ws = wb[index_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    schemes = []
    
    # Data starts at row 2 (index 1) - first row is header
    for i in range(1, len(rows)):
        r = rows[i]
        short_code = clean_str(r[0])
        
        if not short_code:
            continue
        
        # Column 1 might be a formula or actual name
        name = clean_str(r[1]) if len(r) > 1 else None
        
        # If name is a formula or missing, we'll get it from the sheet itself
        schemes.append({
            "scheme_short_code": short_code,
            "scheme_name": name,  # May be None, will be filled from sheet
        })
    
    log.info(f"Found {len(schemes)} schemes in {index_sheet_name} sheet")
    return schemes


def extract_equity_holdings(ws_rows: list[tuple], scheme_info: dict) -> dict | None:
    """
    Extract listed equity holdings from a scheme sheet.
    Returns dict with fund info and holdings, or None if no equity data.
    """
    # Find "Equity & Equity related" section
    equity_section_row = None
    for i, r in enumerate(ws_rows):
        cell_c = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        if cell_c and "EQUITY" in cell_c.upper() and "RELATED" in cell_c.upper():
            equity_section_row = i
            break
    
    if equity_section_row is None:
        return None  # No equity section
    
    # Find "(a) Listed / awaiting listing" subsection
    listed_section_row = None
    for i in range(equity_section_row, min(equity_section_row + 10, len(ws_rows))):
        r = ws_rows[i]
        cell_c = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        if cell_c and ("Listed / awaiting" in cell_c or "Listed/awaiting" in cell_c):
            listed_section_row = i
            break
    
    if listed_section_row is None:
        return None  # No listed equity subsection
    
    # Check if it's NIL
    next_row = ws_rows[listed_section_row + 1] if listed_section_row + 1 < len(ws_rows) else []
    nil_val = clean_str(next_row[COL_MARKET_VALUE]) if len(next_row) > COL_MARKET_VALUE else None
    if nil_val and nil_val.upper() == "NIL":
        return None
    
    # Extract listed equity rows
    holdings = []
    for i in range(listed_section_row + 1, len(ws_rows)):
        r = ws_rows[i]
        
        # Check if we've hit the end (Total row or next category)
        cell_c = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        if cell_c is None:
            continue
        
        cell_c_lower = cell_c.lower().strip()
        
        # End conditions
        if cell_c_lower == "total":
            break
        if cell_c.startswith("(b)") or cell_c.startswith("(c)") or cell_c.startswith("(d)"):
            break
        if "unlisted" in cell_c_lower:
            break
        
        # Validate this is a real data row
        isin = clean_str(r[COL_ISIN]) if len(r) > COL_ISIN else None
        if isin is None or len(isin) != 12:  # ISIN should be 12 characters
            continue
        
        # Parse the data row
        holding = {
            "isin": isin,
            "security_name": cell_c,
            "industry": clean_str(r[COL_INDUSTRY]) if len(r) > COL_INDUSTRY else None,
            "quantity": safe_int(r[COL_QUANTITY]) if len(r) > COL_QUANTITY else None,
            "market_value_lakhs": safe_float(r[COL_MARKET_VALUE]) if len(r) > COL_MARKET_VALUE else None,
            "pct_to_nav": safe_float(r[COL_PCT_NAV]) if len(r) > COL_PCT_NAV else None,
        }
        holdings.append(holding)
    
    if not holdings:
        return None
    
    return {
        "scheme_short_code": scheme_info["scheme_short_code"],
        "scheme_name": scheme_info["scheme_name"],
        "holdings": holdings,
    }


def run_etl(excel_path: str) -> dict:
    """Run ETL process on Nippon India MF Excel file."""
    log.info(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # Parse Index sheet
    schemes = parse_index_sheet(wb)
    
    # Extract report date from first scheme sheet
    report_date = None
    first_sheet = wb[wb.sheetnames[1]]
    first_rows = list(first_sheet.iter_rows(max_row=5, values_only=True))
    for r in first_rows:
        for cell in r:
            if cell and isinstance(cell, str) and "Monthly Portfolio Statement" in cell:
                # Extract date from string like "Monthly Portfolio Statement as on January 31,2026"
                parts = cell.split("as on")
                if len(parts) > 1:
                    date_str = parts[1].strip()
                    report_date = parse_date(date_str)
                    break
        if report_date:
            break
    
    log.info(f"Report date: {report_date}")
    
    # Process each scheme sheet
    all_holdings = []
    funds_with_equity = []
    security_counter = Counter()
    skipped = 0
    errors = 0
    
    for scheme in schemes:
        short_code = scheme["scheme_short_code"]
        
        # Skip Index sheet
        if short_code == "Index":
            continue
        
        try:
            ws = wb[short_code]
            ws_rows = list(ws.iter_rows(values_only=True))
            
            # Get scheme name from B1 if not in index
            if not scheme["scheme_name"] and len(ws_rows) > 0:
                scheme["scheme_name"] = clean_str(ws_rows[0][1]) if len(ws_rows[0]) > 1 else short_code
            
            # Extract equity holdings
            result = extract_equity_holdings(ws_rows, scheme)
            
            if result is None:
                skipped += 1
                log.debug(f"  {short_code}: No equity data")
                continue
            
            num_holdings = len(result["holdings"])
            log.info(f"  {short_code}: Extracted {num_holdings} equity holdings")
            
            # Add to funds with equity
            funds_with_equity.append({
                "scheme_short_code": short_code,
                "scheme_name": result["scheme_name"],
                "scheme_code": short_code,  # Use short code as code
                "holdings_count": num_holdings,
            })
            
            # Process holdings
            for h in result["holdings"]:
                h["scheme_short_code"] = short_code
                all_holdings.append(h)
                security_counter[h["isin"]] += 1
        
        except Exception as e:
            log.error(f"  {short_code}: Error - {e}")
            errors += 1
    
    # Build security master (deduplicated by ISIN)
    security_master = []
    isin_to_info = {}
    
    for h in all_holdings:
        isin = h["isin"]
        if isin not in isin_to_info:
            isin_to_info[isin] = {
                "isin": isin,
                "security_name": h["security_name"],
                "current_industry": h["industry"],
                "current_sector": None,  # Nippon doesn't have separate sector
            }
    
    security_master = list(isin_to_info.values())
    
    # Build output
    output = {
        "metadata": {
            "source_file": os.path.basename(excel_path),
            "report_date": report_date,
            "extraction_date": datetime.now().strftime("%Y-%m-%d"),
            "total_schemes": len(schemes) - 1,  # Exclude Index
            "schemes_with_equity": len(funds_with_equity),
            "schemes_skipped": skipped,
            "errors": errors,
            "total_unique_securities": len(security_master),
            "total_holdings_records": len(all_holdings),
        },
        "amc_master": {
            "amc_name": "Nippon India Mutual Fund",
            "short_code": "NIPPONMF",
        },
        "fund_master": funds_with_equity,
        "security_master": security_master,
        "portfolio_holdings": all_holdings,
    }
    
    # Summary
    log.info("=" * 60)
    log.info("ETL COMPLETE")
    log.info(f"  Schemes processed: {len(funds_with_equity)}/{len(schemes)-1}")
    log.info(f"  Unique securities: {len(security_master)}")
    log.info(f"  Total holdings:    {len(all_holdings)}")
    log.info(f"  Skipped:           {skipped}")
    log.info(f"  Errors:            {errors}")
    log.info("=" * 60)
    
    return output


def main():
    parser = argparse.ArgumentParser(
        description="Extract equity holdings from Nippon India MF portfolio Excel"
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        log.setLevel(logging.DEBUG)
    
    if not os.path.exists(args.excel_file):
        log.error(f"File not found: {args.excel_file}")
        sys.exit(1)
    
    # Run ETL
    result = run_etl(args.excel_file)
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        # Auto-generate based on report date
        report_date = result["metadata"]["report_date"]
        if report_date:
            date_suffix = report_date.replace("-", "")[:6]  # YYYYMM
        else:
            date_suffix = datetime.now().strftime("%Y%m")
        
        output_dir = Path("data/processed")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"nippon_equity_holdings_{date_suffix}.json"
    
    # Write JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    
    log.info(f"\n[OK] Output written to: {output_path}")


if __name__ == "__main__":
    main()
