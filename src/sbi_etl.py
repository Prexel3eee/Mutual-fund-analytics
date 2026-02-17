"""
SBI Mutual Fund ETL Pipeline
Extracts equity holdings data from SBI MF monthly portfolio Excel files.

Output: JSON file structured to match the database schema:
  - amc_master
  - fund_master (schemes)
  - security_master (unique securities)
  - portfolio_holdings (per-scheme equity positions)

Usage:
    python src/sbi_etl.py <excel_file> [--output <output_file>]
    python src/sbi_etl.py January.xlsx
    python src/sbi_etl.py January.xlsx --output data/processed/sbi_202601.json
"""

import argparse
import json
import logging
import os
import sys
from datetime import datetime
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("sbi_etl")

# ---------------------------------------------------------------------------
# Constants – column indices (0-based) inside each scheme sheet
# ---------------------------------------------------------------------------
COL_SECURITY_CODE = 1   # Column B  – internal SBI security code
COL_NAME          = 2   # Column C  – Name of the Instrument / Issuer
COL_ISIN          = 3   # Column D  – ISIN
COL_INDUSTRY      = 4   # Column E  – Rating / Industry
COL_QUANTITY      = 5   # Column F  – Quantity
COL_MARKET_VALUE  = 6   # Column G  – Market value (Rs. in Lakhs)
COL_PCT_AUM       = 7   # Column H  – % to AUM

# Metadata rows (1-based, but we use 0-based for list indexing)
ROW_FUND_NAME   = 1  # index 1 → Row 2
ROW_SCHEME_NAME = 2  # index 2 → Row 3
ROW_DATE        = 3  # index 3 → Row 4
ROW_HEADERS     = 5  # index 5 → Row 6

# Category markers (appear in column C / index 2)
MARKER_EQUITY_SECTION = "EQUITY & EQUITY RELATED"
MARKER_LISTED = "a) Listed/awaiting listing on Stock Exchanges"

# Markers that signal end of listed equity section
END_MARKERS = {"total", "b) unlisted", "b)", "c)", "d)"}


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def safe_float(val: Any) -> float | None:
    """Convert a cell value to float, returning None for non-numeric."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "NIL", "N/A", "#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def safe_int(val: Any) -> int | None:
    """Convert a cell value to int (quantity), returning None for non-numeric."""
    f = safe_float(val)
    if f is None:
        return None
    return int(f)


def clean_str(val: Any) -> str | None:
    """Strip and clean a string cell value."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_date(val: Any) -> str | None:
    """Parse a datetime or date string into YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # return raw string if unparseable


# ---------------------------------------------------------------------------
# Index Sheet Parser
# ---------------------------------------------------------------------------

def parse_index_sheet(wb: openpyxl.Workbook) -> list[dict]:
    """Parse the Index sheet and return list of scheme metadata dicts."""
    ws = wb["Index"]
    rows = list(ws.iter_rows(values_only=True))
    schemes = []
    
    # Data starts at row 4 (index 3) after header row at row 3 (index 2)
    for i in range(3, len(rows)):
        r = rows[i]
        code = clean_str(r[0])
        short_code = clean_str(r[1])
        name = clean_str(r[2])
        
        if code and short_code:
            schemes.append({
                "scheme_code": code,
                "scheme_short_code": short_code,
                "scheme_name": name or "",
            })
    
    return schemes


# ---------------------------------------------------------------------------
# Scheme Sheet Parser – Equity Extraction
# ---------------------------------------------------------------------------

def extract_equity_holdings(ws_rows: list[tuple], scheme_info: dict) -> dict | None:
    """
    Extract equity holdings from a single scheme sheet.
    
    Returns a dict with:
      - fund metadata (scheme name, report date)
      - list of equity holdings
    Or None if no equity data found.
    """
    if len(ws_rows) < 10:
        return None
    
    # --- Extract metadata ---
    report_date = parse_date(ws_rows[ROW_DATE][3]) if len(ws_rows) > ROW_DATE else None
    scheme_name_full = clean_str(ws_rows[ROW_SCHEME_NAME][3]) if len(ws_rows) > ROW_SCHEME_NAME else None
    scheme_code_from_sheet = clean_str(ws_rows[ROW_FUND_NAME][3]) if len(ws_rows) > ROW_FUND_NAME else None
    
    # --- Locate equity section ---
    equity_section_row = None
    listed_section_row = None
    
    for i, r in enumerate(ws_rows):
        cell_c = clean_str(r[COL_NAME])
        if cell_c is None:
            continue
        
        cell_upper = cell_c.upper()
        
        if "EQUITY" in cell_upper and "RELATED" in cell_upper:
            equity_section_row = i
        
        if equity_section_row is not None and "LISTED/AWAITING" in cell_upper and "STOCK EXCHANGE" in cell_upper:
            listed_section_row = i
            break
    
    if listed_section_row is None:
        log.debug(f"  No listed equity section found in {scheme_info['scheme_short_code']}")
        return None
    
    # Check if it's NIL
    nil_val = clean_str(ws_rows[listed_section_row][COL_MARKET_VALUE])
    if nil_val and nil_val.upper() == "NIL":
        log.debug(f"  Equity section is NIL in {scheme_info['scheme_short_code']}")
        return None
    
    # --- Extract listed equity rows ---
    holdings = []
    
    for i in range(listed_section_row + 1, len(ws_rows)):
        r = ws_rows[i]
        cell_c = clean_str(r[COL_NAME])
        
        if cell_c is None:
            continue
        
        cell_lower = cell_c.lower().strip()
        
        # Check if we've hit the end of listed equity
        if cell_lower == "total":
            break
        if cell_lower.startswith("b)") or cell_lower.startswith("c)") or cell_lower.startswith("d)"):
            break
        if any(marker in cell_lower for marker in ["debt instruments", "money market", "others", "unlisted"]):
            break
        
        # Validate this is a real data row (must have security code and ISIN)
        security_code = clean_str(r[COL_SECURITY_CODE])
        isin = clean_str(r[COL_ISIN])
        
        if security_code is None or isin is None:
            continue
        
        # Parse the data row
        holding = {
            "security_code": security_code,
            "security_name": cell_c,
            "isin": isin,
            "industry": clean_str(r[COL_INDUSTRY]),
            "quantity": safe_int(r[COL_QUANTITY]),
            "market_value_lakhs": safe_float(r[COL_MARKET_VALUE]),
            "pct_to_aum": safe_float(r[COL_PCT_AUM]),
        }
        
        holdings.append(holding)
    
    if not holdings:
        return None
    
    return {
        "scheme_code": scheme_info["scheme_code"],
        "scheme_short_code": scheme_info["scheme_short_code"],
        "scheme_name": scheme_name_full or scheme_info["scheme_name"],
        "report_date": report_date,
        "holdings_count": len(holdings),
        "holdings": holdings,
    }


# ---------------------------------------------------------------------------
# Main ETL Pipeline
# ---------------------------------------------------------------------------

def run_etl(excel_path: str) -> dict:
    """
    Run the full ETL pipeline on an SBI MF portfolio Excel file.
    
    Returns a structured dict matching the database schema.
    """
    log.info(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    log.info(f"Total sheets: {len(wb.sheetnames)}")
    
    # --- Step 1: Parse Index ---
    schemes = parse_index_sheet(wb)
    log.info(f"Parsed Index sheet: {len(schemes)} schemes found")
    
    # --- Step 2: Process each scheme sheet ---
    processed_funds = []
    all_securities = {}  # ISIN -> security info (deduplicated)
    all_holdings = []    # flat list for portfolio_holdings
    skipped_count = 0
    error_count = 0
    
    for scheme in schemes:
        short_code = scheme["scheme_short_code"]
        
        if short_code not in wb.sheetnames:
            log.warning(f"  Sheet '{short_code}' not found in workbook, skipping")
            skipped_count += 1
            continue
        
        ws = wb[short_code]
        ws_rows = list(ws.iter_rows(values_only=True))
        
        try:
            result = extract_equity_holdings(ws_rows, scheme)
        except Exception as e:
            log.error(f"  Error processing sheet '{short_code}': {e}")
            error_count += 1
            continue
        
        if result is None:
            log.debug(f"  {short_code}: No equity data, skipping")
            skipped_count += 1
            continue
        
        log.info(f"  {short_code}: Extracted {result['holdings_count']} equity holdings")
        
        # Build fund_master entry
        fund_entry = {
            "scheme_code": result["scheme_code"],
            "scheme_short_code": result["scheme_short_code"],
            "scheme_name": result["scheme_name"],
            "amc_id": "SBI",
            "report_date": result["report_date"],
            "holdings_count": result["holdings_count"],
        }
        processed_funds.append(fund_entry)
        
        # Process each holding
        for h in result["holdings"]:
            # Add to security_master (deduplicated by ISIN)
            isin = h["isin"]
            if isin not in all_securities:
                all_securities[isin] = {
                    "isin": isin,
                    "security_name": h["security_name"],
                    "asset_class": "Equity",
                    "current_sector": h["industry"],
                    "current_industry": h["industry"],
                }
            
            # Add to portfolio_holdings
            all_holdings.append({
                "scheme_code": result["scheme_code"],
                "scheme_short_code": result["scheme_short_code"],
                "isin": isin,
                "security_name": h["security_name"],
                "industry": h["industry"],
                "quantity": h["quantity"],
                "market_value_lakhs": h["market_value_lakhs"],
                "pct_to_aum": h["pct_to_aum"],
                "report_date": result["report_date"],
            })
    
    wb.close()
    
    # --- Step 3: Build final JSON output ---
    report_date = processed_funds[0]["report_date"] if processed_funds else None
    
    output = {
        "metadata": {
            "source_file": os.path.basename(excel_path),
            "amc": "SBI Mutual Fund",
            "report_date": report_date,
            "extracted_at": datetime.now().isoformat(),
            "total_schemes_in_file": len(schemes),
            "schemes_with_equity": len(processed_funds),
            "schemes_skipped": skipped_count,
            "schemes_errored": error_count,
            "total_unique_securities": len(all_securities),
            "total_holdings_records": len(all_holdings),
        },
        "amc_master": {
            "amc_name": "SBI Mutual Fund",
            "short_code": "SBI",
        },
        "fund_master": processed_funds,
        "security_master": list(all_securities.values()),
        "portfolio_holdings": all_holdings,
    }
    
    # --- Summary stats ---
    log.info("=" * 60)
    log.info(f"ETL COMPLETE")
    log.info(f"  Schemes processed: {len(processed_funds)}/{len(schemes)}")
    log.info(f"  Unique securities: {len(all_securities)}")
    log.info(f"  Total holdings:    {len(all_holdings)}")
    log.info(f"  Skipped:           {skipped_count}")
    log.info(f"  Errors:            {error_count}")
    log.info("=" * 60)
    
    return output


def main():
    parser = argparse.ArgumentParser(
        description="SBI Mutual Fund ETL – Extract equity holdings from portfolio Excel files"
    )
    parser.add_argument("excel_file", help="Path to the SBI MF portfolio Excel file")
    parser.add_argument(
        "--output", "-o",
        help="Output JSON file path (default: sbi_equity_holdings_YYYYMM.json)",
        default=None,
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable debug logging",
    )
    
    args = parser.parse_args()
    
    if args.verbose:
        log.setLevel(logging.DEBUG)
    
    if not os.path.exists(args.excel_file):
        log.error(f"File not found: {args.excel_file}")
        sys.exit(1)
    
    # Run ETL
    result = run_etl(args.excel_file)
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        report_date = result["metadata"]["report_date"]
        if report_date:
            dt = datetime.strptime(report_date, "%Y-%m-%d")
            date_suffix = dt.strftime("%Y%m")
        else:
            date_suffix = "unknown"
        output_path = f"sbi_equity_holdings_{date_suffix}.json"
    
    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    
    # Write JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    
    log.info(f"Output written to: {output_path}")
    
    # Print quick summary to stdout
    print(f"\n[OK] Extracted {len(result['fund_master'])} equity schemes, "
          f"{len(result['security_master'])} unique securities, "
          f"{len(result['portfolio_holdings'])} total holdings")
    print(f"[->] Output: {output_path}")


if __name__ == "__main__":
    main()
