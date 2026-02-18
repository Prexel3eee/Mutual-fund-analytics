"""
ETL script for Edelweiss Mutual Fund monthly portfolio Excel files.
Extracts Listed Equity holdings and outputs JSON matching database schema.
"""

import argparse
import json
import logging
import os
import sys
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

# Column indices (0-indexed) based on Row 4 headers:
# Name of the Instrument | ISIN | Rating/Industry | Quantity | Market/Fair Value | % to Net Assets | YIELD
COL_NAME = 0
COL_ISIN = 1
COL_INDUSTRY = 2
COL_QUANTITY = 3
COL_MARKET_VALUE = 4
COL_PCT_NAV = 5

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("edelweiss_etl")


def safe_float(val: Any) -> float | None:
    """Convert value to float, return None if invalid."""
    if val is None:
        return None
    try:
        v = float(val)
        return v if v != 0 else None
    except (ValueError, TypeError):
        return None


def safe_int(val: Any) -> int | None:
    """Convert value to int, return None if invalid."""
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def clean_str(val: Any) -> str | None:
    """Clean string value, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_date(val: Any) -> str | None:
    """Parse date from header text like 'PORTFOLIO STATEMENT as on 31 Jan 2026'."""
    if val is None:
        return None
    s = str(val).strip()

    # Pattern: "as on DD Mon YYYY" or "as on DD-Mon-YYYY"
    m = re.search(r'as\s+on\s+(\d{1,2})\s+(\w+)\s+(\d{4})', s, re.IGNORECASE)
    if m:
        day, mon, year = m.group(1), m.group(2), m.group(3)
        for fmt in ("%d %b %Y", "%d %B %Y"):
            try:
                return datetime.strptime(f"{day} {mon} {year}", fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue

    # Pattern: "DD-Mon-YYYY"
    m = re.search(r'(\d{1,2})-(\w{3})-(\d{4})', s)
    if m:
        try:
            return datetime.strptime(m.group(0), "%d-%b-%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass

    return None


def parse_index_sheet(wb: openpyxl.Workbook) -> dict[str, str]:
    """Parse Index sheet to map sheet codes to scheme names.

    Index sheet structure:
    - Row 1: AMC name
    - Row 2: Portfolio date
    - Row 3: Headers (Fund Id, Fund Desc, ...)
    - Row 4+: Data rows

    Returns:
        dict mapping sheet code (Fund Id) to scheme name (Fund Desc)
    """
    if "Index" not in wb.sheetnames:
        log.warning("No 'Index' sheet found")
        return {}

    ws = wb["Index"]
    rows = list(ws.iter_rows(values_only=True))

    scheme_map = {}

    # Data starts from row 4 (index 3), after header row
    for i, r in enumerate(rows[3:], start=4):
        fund_id = clean_str(r[0]) if len(r) > 0 else None
        fund_desc = clean_str(r[1]) if len(r) > 1 else None

        if fund_id and fund_desc:
            scheme_map[fund_id] = fund_desc
            log.debug(f"  {fund_id} -> {fund_desc}")

    log.info(f"Parsed Index sheet: {len(scheme_map)} schemes")
    return scheme_map


def extract_equity_holdings(ws_rows: list[tuple], scheme_info: dict) -> dict | None:
    """Extract listed equity holdings from a scheme sheet.

    Returns dict with fund info and holdings, or None if no equity data.
    """
    scheme_name = scheme_info.get("scheme_name", "Unknown")

    # Find equity section and listed subsection
    equity_row = None
    listed_row = None

    for i, r in enumerate(ws_rows):
        cell0 = clean_str(r[0]) if len(r) > 0 else None
        if not cell0:
            continue

        cl = cell0.lower()

        if 'equity' in cl and 'related' in cl and equity_row is None:
            equity_row = i
        elif '(a)' in cl and 'listed' in cl.lower():
            listed_row = i
        elif cl.startswith('(a)') and 'awaiting' in cl.lower():
            listed_row = i

    if equity_row is None:
        return None  # No equity section

    # If no explicit listed section found, start from equity row + 1
    start_row = (listed_row if listed_row is not None else equity_row) + 1

    holdings = []
    isin_map = {}  # For deduplication

    for i in range(start_row, len(ws_rows)):
        r = ws_rows[i]

        # Check for end markers in Col 0
        cell0 = clean_str(r[0]) if len(r) > 0 else None
        if cell0:
            cl = cell0.lower().strip()
            if cl in ('sub total', 'total', 'sub-total'):
                break
            if cl.startswith('(b)') or cl.startswith('(c)'):
                break

        # Extract data
        name = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        isin = clean_str(r[COL_ISIN]) if len(r) > COL_ISIN else None
        industry = clean_str(r[COL_INDUSTRY]) if len(r) > COL_INDUSTRY else None
        quantity = safe_float(r[COL_QUANTITY]) if len(r) > COL_QUANTITY else None
        market_value = safe_float(r[COL_MARKET_VALUE]) if len(r) > COL_MARKET_VALUE else None
        pct_nav = safe_float(r[COL_PCT_NAV]) if len(r) > COL_PCT_NAV else None

        # Must have ISIN and name
        if not isin or not name:
            continue

        # ISIN validation
        if len(isin) != 12 or not isin.startswith("INE"):
            log.debug(f"  Skipping invalid ISIN: {isin}")
            continue

        # Aggregate duplicates
        if isin in isin_map:
            idx = isin_map[isin]
            existing = holdings[idx]
            if quantity and existing.get("quantity"):
                existing["quantity"] += quantity
            if market_value and existing.get("market_value_lakhs"):
                existing["market_value_lakhs"] += market_value
            continue

        holding = {
            "security_name": name,
            "isin": isin,
            "industry": industry,
            "quantity": safe_int(quantity) if quantity else None,
            "market_value_lakhs": round(market_value, 4) if market_value else None,
            "pct_to_nav": round(pct_nav, 4) if pct_nav else None,
        }

        isin_map[isin] = len(holdings)
        holdings.append(holding)

    if not holdings:
        return None

    return {
        "scheme_name": scheme_name,
        "scheme_code": scheme_info.get("sheet_code", ""),
        "holdings": holdings,
    }


def run_etl(excel_path: str, date_override: str | None = None) -> dict:
    """Run the ETL process on an Edelweiss Mutual Fund portfolio Excel file."""
    log.info(f"Processing: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Parse index sheet
    scheme_map = parse_index_sheet(wb)

    # Extract report date
    report_date = None

    if date_override:
        report_date = date_override
        log.info(f"Using provided report date: {report_date}")
    else:
        # Try from Index sheet Row 2
        if "Index" in wb.sheetnames:
            ws = wb["Index"]
            rows = list(ws.iter_rows(max_row=5, values_only=True))
            for r in rows:
                for cell in r:
                    if cell:
                        parsed = parse_date(cell)
                        if parsed:
                            report_date = parsed
                            break
                if report_date:
                    break

    if not report_date:
        log.warning("Could not extract report date, using current date")
        report_date = datetime.now().strftime("%Y-%m-%d")

    log.info(f"Report date: {report_date}")

    # Process each scheme sheet
    all_funds = []
    all_securities = {}
    all_holdings = []
    skipped = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "index":
            continue

        scheme_name = scheme_map.get(sheet_name, f"Unknown ({sheet_name})")

        try:
            ws = wb[sheet_name]
            ws_rows = list(ws.iter_rows(values_only=True))

            result = extract_equity_holdings(ws_rows, {
                "scheme_name": scheme_name,
                "sheet_code": sheet_name,
            })

            if not result:
                skipped += 1
                continue

            holdings = result["holdings"]
            fund_id = f"EDELWEISS_{sheet_name}"

            fund_record = {
                "fund_id": fund_id,
                "scheme_name": result["scheme_name"],
                "scheme_code": sheet_name,
                "scheme_short_code": sheet_name,  # Required by load_to_postgres.py
            }
            all_funds.append(fund_record)

            for h in holdings:
                isin = h["isin"]
                if isin not in all_securities:
                    all_securities[isin] = {
                        "isin": isin,
                        "security_name": h["security_name"],
                        "industry": h.get("industry"),
                    }

                all_holdings.append({
                    "fund_id": fund_id,
                    "scheme_short_code": sheet_name,  # Required by load_to_postgres.py
                    "isin": isin,
                    "report_date": report_date,
                    "quantity": h.get("quantity"),
                    "market_value_lakhs": h.get("market_value_lakhs"),
                    "pct_to_nav": h.get("pct_to_nav"),
                })

            log.info(f"  ✓ {sheet_name}: {len(holdings)} holdings")

        except Exception as e:
            log.warning(f"  ✗ {sheet_name}: {e}")
            skipped += 1

    log.info(f"  Schemes: {len(all_funds)}, Skipped: {skipped}")
    log.info(f"  Securities: {len(all_securities)}")
    log.info(f"  Total holdings: {len(all_holdings)}")

    return {
        "metadata": {
            "source_file": os.path.basename(excel_path),
            "amc": "Edelweiss Mutual Fund",
            "report_date": report_date,
            "total_schemes": len(wb.sheetnames) - 1,
            "schemes_with_equity": len(all_funds),
            "total_unique_securities": len(all_securities),
            "total_holdings_records": len(all_holdings),
        },
        "amc_master": {
            "amc_name": "Edelweiss Mutual Fund",
            "short_code": "EDELWEISS",
        },
        "fund_master": all_funds,
        "security_master": list(all_securities.values()),
        "portfolio_holdings": all_holdings,
    }


def main():
    parser = argparse.ArgumentParser(
        description="ETL for Edelweiss Mutual Fund portfolio Excel files"
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--date", "-d", help="Override report date (YYYY-MM-DD)")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Run ETL
    data = run_etl(args.excel_file, date_override=args.date)

    # Determine output path
    if args.output:
        out_path = args.output
    else:
        stem = Path(args.excel_file).stem.lower()
        out_path = f"data/processed/edelweiss/edelweiss_equity_holdings_{stem}.json"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    log.info(f"Written to: {out_path}")


if __name__ == "__main__":
    main()
