"""
ETL script for Axis Mutual Fund monthly portfolio Excel files.
Extracts Listed Equity holdings and outputs JSON matching database schema.
"""

import argparse
import json
import logging
import os
import sys
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

# Column indices for Axis Mutual Fund (0-indexed)
COL_NAME = 1            # Security Name  
COL_ISIN = 2            # ISIN Code
COL_INDUSTRY = 3        # Industry/Rating
COL_QUANTITY = 4        # Quantity
COL_MARKET_VALUE = 5    # Market Value (Rs. in Lacs)
COL_PCT_NAV = 6         # % to Net Assets

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("axis_etl")


def safe_float(val: Any) -> float | None:
    """Convert value to float, return None if invalid."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "NIL", "N/A", "#", "-", "Total", "Sub Total"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def safe_int(val: Any) -> int | None:
    """Convert value to int, return None if invalid."""
    f = safe_float(val)
    return int(f) if f is not None else None


def clean_str(val: Any) -> str | None:
    """Clean string value, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in ("", "NIL", "N/A", "#", "-") else None


def parse_date(val: Any) -> str | None:
    """Parse date to YYYY-MM-DD format."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    
    # Clean up common issues
    s = s.replace(",", ", ") # Ensure space after comma
    s = re.sub(r'\s+', ' ', s) # Collapse multiple spaces
    
    # Try common formats
    formats = (
        "%Y-%m-%d %H:%M:%S", 
        "%Y-%m-%d", 
        "%d-%m-%Y", 
        "%d/%m/%Y", 
        "%B %d, %Y",  # January 31, 2026
        "%B %d,%Y",   # January 31,2026
        "%d-%b-%Y",   # 31-Jan-2026
        "%d %B %Y",   # 31 January 2026
    )
    
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
            
    # Try with Title Case if simpler checking failed (fixes "january" -> "January")
    s_title = s.title()
    for fmt in formats:
        try:
            return datetime.strptime(s_title, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
            
    return s


def parse_index_sheet(wb: openpyxl.Workbook) -> dict[str, str]:
    """
    Parse 'Index' sheet to get short name → scheme name mapping.
    Returns dict of {short_name: full_name}.
    """
    if "Index" not in wb.sheetnames:
        log.warning("No 'Index' sheet found")
        return {}
    
    ws = wb["Index"]
    rows = list(ws.iter_rows(values_only=True))
    
    mapping = {}
    # Header at row 1: Sr No. | Short Name | Scheme Name
    # Data starts from row 2 (index 1)
    for i in range(1, len(rows)):
        r = rows[i]
        if len(r) < 3:
            continue
        
        short_name = clean_str(r[1])
        scheme_name = clean_str(r[2])
        
        if short_name and scheme_name:
            mapping[short_name] = scheme_name
    
    log.info(f"Found {len(mapping)} scheme mappings in Index sheet")
    return mapping


def extract_equity_holdings(ws_rows: list[tuple], scheme_info: dict) -> dict | None:
    """
    Extract listed equity holdings from a scheme sheet.
    Returns dict with fund info and holdings, or None if no equity data.
    """
    # Find "Equity & Equity related" section - Axis puts this in column 1
    equity_section_row = None
    for i, r in enumerate(ws_rows):
        cell = clean_str(r[1]) if len(r) > 1 else None
        if cell and "EQUITY" in cell.upper() and "RELATED" in cell.upper():
            equity_section_row = i
            break
    
    if equity_section_row is None:
        return None  # No equity section
    
    # Find "(a) Listed" subsection - also in column 1
    listed_section_row = None
    for i in range(equity_section_row, min(equity_section_row + 10, len(ws_rows))):
        r = ws_rows[i]
        cell = clean_str(r[1]) if len(r) > 1 else None
        if cell and ("(a)" in cell and ("Listed" in cell or "listed" in cell)):
            listed_section_row = i
            break
    
    if listed_section_row is None:
        return None  # No listed equity subsection
    
    # Extract listed equity rows
    holdings = []
    for i in range(listed_section_row + 1, len(ws_rows)):
        r = ws_rows[i]
        
        # Check if we've hit the end
        cell_name = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        
        if cell_name is None:
            continue  # Skip empty rows
        
        # Validate this is a real data row first
        isin = clean_str(r[COL_ISIN]) if len(r) > COL_ISIN else None
        
        # End conditions - check column 1 for subsection markers
        # Only check end conditions if we don't have a valid ISIN, 
        # OR if the name is strictly "Total" or "Sub Total"
        cell_col1 = clean_str(r[1]) if len(r) > 1 else None
        
        if cell_col1:
            c_lower = cell_col1.lower().strip()
            # Strict checks for Total/Sub Total to avoid matching companies like "Adani Total Gas"
            is_end_marker = (
                c_lower == "total" or 
                c_lower == "sub total" or 
                c_lower.startswith("(b)") or 
                c_lower.startswith("unlisted")
            )
            
            # If it's an end marker and (no ISIN or looks like a header), stop
            if is_end_marker:
                # Double check: if it has a valid ISIN, it's probably a company name that looks like a marker
                # But headers/totals usually don't have valid ISINs (Axis totals don't)
                if not isin or len(isin) != 12:
                    break
        
        if isin is None or len(isin) != 12:  # ISIN should be 12 characters
            continue
        
        # Parse the data row
        holding = {
            "isin": isin,
            "security_name": cell_name,
            "industry": clean_str(r[COL_INDUSTRY]) if len(r) > COL_INDUSTRY else None,
            "quantity": safe_int(r[COL_QUANTITY]) if len(r) > COL_QUANTITY else None,
            "market_value_lakhs": safe_float(r[COL_MARKET_VALUE]) if len(r) > COL_MARKET_VALUE else None,
            "pct_to_nav": safe_float(r[COL_PCT_NAV]) if len(r) > COL_PCT_NAV else None,
        }
        holdings.append(holding)
    
    if not holdings:
        return None
    
    # Aggregate duplicate ISINs within the scheme
    aggregated_holdings = {}
    for h in holdings:
        isin = h["isin"]
        if isin in aggregated_holdings:
            existing = aggregated_holdings[isin]
            # Sum quantity
            if h["quantity"] is not None:
                existing["quantity"] = (existing["quantity"] or 0) + h["quantity"]
            # Sum market value
            if h["market_value_lakhs"] is not None:
                existing["market_value_lakhs"] = (existing["market_value_lakhs"] or 0) + h["market_value_lakhs"]
            # Sum pct
            if h["pct_to_nav"] is not None:
                existing["pct_to_nav"] = (existing["pct_to_nav"] or 0) + h["pct_to_nav"]
        else:
            aggregated_holdings[isin] = h
            
    final_holdings = list(aggregated_holdings.values())
    
    return {
        "scheme_short_code": scheme_info["scheme_short_code"],
        "scheme_name": scheme_info["scheme_name"],
        "holdings": final_holdings,
    }


def run_etl(excel_path: str) -> dict:
    """Run ETL process on Axis Mutual Fund Excel file."""
    log.info(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # Parse Index sheet for short name mapping
    scheme_mapping = parse_index_sheet(wb)
    
    # Extract report date from first data sheet header
    report_date = None
    for sheet_name in wb.sheetnames[1:]:  # Skip Index
        ws = wb[sheet_name]
        first_rows = list(ws.iter_rows(max_row=3, values_only=True))
        for r in first_rows:
            for cell in r:
                if cell and isinstance(cell, str):
                    # Case-insensitive search but extraction preserves case
                    if "as on" in cell.lower():
                        # Use regex to split case-insensitive
                        parts = re.split(r'as on', cell, flags=re.IGNORECASE)
                        if len(parts) > 1:
                            date_str = parts[1].strip()
                            report_date = parse_date(date_str)
                            break
            if report_date:
                break
        if report_date:
            break
    
    log.info(f"Report date: {report_date}")
    
    # Process each scheme sheet (skip "Index")
    all_holdings = []
    funds_with_equity = []
    security_counter = Counter()
    skipped = 0
    errors = 0
    
    for sheet_name in wb.sheetnames:
        # Skip Index sheet
        if sheet_name == "Index":
            continue
        
        try:
            ws = wb[sheet_name]
            ws_rows = list(ws.iter_rows(values_only=True))
            
            # Get full scheme name from mapping, or use sheet name
            full_name = scheme_mapping.get(sheet_name, sheet_name)
            
            scheme_info = {
                "scheme_short_code": sheet_name,
                "scheme_name": full_name,
            }
            
            # Extract equity holdings
            result = extract_equity_holdings(ws_rows, scheme_info)
            
            if result is None:
                skipped += 1
                log.debug(f"  {sheet_name}: No equity data")
                continue
            
            num_holdings = len(result["holdings"])
            log.info(f"  {sheet_name}: Extracted {num_holdings} equity holdings")
            
            # Add to funds with equity
            funds_with_equity.append({
                "scheme_short_code": sheet_name,
                "scheme_name": full_name,
                "scheme_code": sheet_name,
                "holdings_count": num_holdings,
            })
            
            # Process holdings
            for h in result["holdings"]:
                h["scheme_short_code"] = sheet_name
                all_holdings.append(h)
                security_counter[h["isin"]] += 1
        
        except Exception as e:
            log.error(f"  {sheet_name}: Error - {e}")
            errors += 1
    
    # Build security master (deduplicated by ISIN)
    security_master = []
    isin_to_info = {}
    
    for h in all_holdings:
        isin = h["isin"]
        if isin not in isin_to_info:
            isin_to_info[isin] = {
                "isin": isin,
                "security_name": h["security_name"],
                "current_industry": h["industry"],
                "current_sector": None,  # Axis doesn't have separate sector
            }
    
    security_master = list(isin_to_info.values())
    
    # Build output
    output = {
        "metadata": {
            "source_file": os.path.basename(excel_path),
            "report_date": report_date,
            "extraction_date": datetime.now().strftime("%Y-%m-%d"),
            "total_schemes": len(wb.sheetnames) - 1,  # Exclude Index
            "schemes_with_equity": len(funds_with_equity),
            "schemes_skipped": skipped,
            "errors": errors,
            "total_unique_securities": len(security_master),
            "total_holdings_records": len(all_holdings),
        },
        "amc_master": {
            "amc_name": "Axis Mutual Fund",
            "short_code": "AXIS",
        },
        "fund_master": funds_with_equity,
        "security_master": security_master,
        "portfolio_holdings": all_holdings,
    }
    
    # Summary
    log.info("=" * 60)
    log.info("ETL COMPLETE")
    log.info(f"  Schemes processed: {len(funds_with_equity)}/{len(wb.sheetnames)-1}")
    log.info(f"  Unique securities: {len(security_master)}")
    log.info(f"  Total holdings:    {len(all_holdings)}")
    log.info(f"  Skipped:           {skipped}")
    log.info(f"  Errors:            {errors}")
    log.info("=" * 60)
    
    return output


def main():
    parser = argparse.ArgumentParser(
        description="Extract equity holdings from Axis MF portfolio Excel"
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        log.setLevel(logging.DEBUG)
    
    if not os.path.exists(args.excel_file):
        log.error(f"File not found: {args.excel_file}")
        sys.exit(1)
    
    # Run ETL
    result = run_etl(args.excel_file)
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        # Auto-generate based on report date
        report_date = result["metadata"]["report_date"]
        if report_date:
            date_suffix = report_date.replace("-", "")[:6]  # YYYYMM
        else:
            date_suffix = datetime.now().strftime("%Y%m")
        
        output_dir = Path("data/processed")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"axis_equity_holdings_{date_suffix}.json"
    
    # Write JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    
    log.info(f"\n[OK] Output written to: {output_path}")


if __name__ == "__main__":
    main()
