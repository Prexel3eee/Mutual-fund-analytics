"""
ETL script for Kotak Mahindra Mutual Fund monthly portfolio Excel files.
Extracts Listed Equity holdings and outputs JSON matching database schema.
"""

import argparse
import json
import logging
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

# Column indices for Kotak Mahindra (0-indexed)
COL_NAME = 2            # Security Name
COL_ISIN = 3            # ISIN Code
COL_INDUSTRY = 4        # Industry
COL_YIELD = 5           # Yield (optional)
COL_QUANTITY = 6        # Quantity
COL_MARKET_VALUE = 7    # Market Value (Rs. in Lacs)
COL_PCT_NAV = 8         # % to Net Assets

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("kotak_etl")


def safe_float(val: Any) -> float | None:
    """Convert value to float, return None if invalid."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "NIL", "N/A", "#", "-", "Total"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def safe_int(val: Any) -> int | None:
    """Convert value to int, return None if invalid."""
    f = safe_float(val)
    return int(f) if f is not None else None


def clean_str(val: Any) -> str | None:
    """Clean string value, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in ("", "NIL", "N/A", "#", "-") else None


def parse_date(val: Any) -> str | None:
    """Parse date to YYYY-MM-DD format."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%B %d,%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def parse_scheme_sheet(wb: openpyxl.Workbook) -> dict[str, str]:
    """
    Parse 'Scheme' sheet to get abbreviation → scheme name mapping.
    Returns dict of {abbr: full_name}.
    """
    if "Scheme" not in wb.sheetnames:
        log.warning("No 'Scheme' sheet found")
        return {}
    
    ws = wb["Scheme"]
    rows = list(ws.iter_rows(values_only=True))
    
    mapping = {}
    # Header is at row 2 (index 1): Abbreviations | Scheme Name
    # Data starts from row 3 (index 2)
    for i in range(2, len(rows)):
        r = rows[i]
        abbr = clean_str(r[0])
        name = clean_str(r[1])
        
        if abbr and name:
            mapping[abbr] = name
    
    log.info(f"Found {len(mapping)} scheme mappings in Scheme sheet")
    return mapping


def extract_equity_holdings(ws_rows: list[tuple], scheme_info: dict) -> dict | None:
    """
    Extract listed equity holdings from a scheme sheet.
    Returns dict with fund info and holdings, or None if no equity data.
    """
    # Find "Equity & Equity related" section
    equity_section_row = None
    for i, r in enumerate(ws_rows):
        cell = clean_str(r[0]) if len(r) > 0 else None
        if cell and "EQUITY" in cell.upper() and "RELATED" in cell.upper():
            equity_section_row = i
            break
    
    if equity_section_row is None:
        return None  # No equity section
    
    # Find "Listed/Awaiting listing" subsection
    listed_section_row = None
    for i in range(equity_section_row, min(equity_section_row + 10, len(ws_rows))):
        r = ws_rows[i]
        cell = clean_str(r[1]) if len(r) > 1 else None
        if cell and ("Listed/Awaiting" in cell or "Listed/awaiting" in cell):
            listed_section_row = i
            break
    
    if listed_section_row is None:
        return None # No listed equity subsection
    
    # Extract listed equity rows
    holdings = []
    for i in range(listed_section_row + 1, len(ws_rows)):
        r = ws_rows[i]
        
        # Check if we've hit the end (Total row, Next category, or empty rows)
        cell_name = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        cell_industry = clean_str(r[COL_INDUSTRY]) if len(r) > COL_INDUSTRY else None
        
        if cell_name is None and cell_industry is None:
            continue  # Skip empty rows
        
        # End conditions
        if cell_industry and cell_industry.lower().strip() == "total":
            break
        if cell_name and ("unlisted" in cell_name.lower() or "awaiting listing" in cell_name.lower()):
            break
        
        # Validate this is a real data row     
        isin = clean_str(r[COL_ISIN]) if len(r) > COL_ISIN else None
        if isin is None or len(isin) != 12:  # ISIN should be 12 characters
            continue
        
        # Parse the data row
        holding = {
            "isin": isin,
            "security_name": cell_name,
            "industry": cell_industry,
            "quantity": safe_int(r[COL_QUANTITY]) if len(r) > COL_QUANTITY else None,
            "market_value_lakhs": safe_float(r[COL_MARKET_VALUE]) if len(r) > COL_MARKET_VALUE else None,
            "pct_to_nav": safe_float(r[COL_PCT_NAV]) if len(r) > COL_PCT_NAV else None,
        }
        holdings.append(holding)
    
    if not holdings:
        return None
    
    return {
        "scheme_short_code": scheme_info["scheme_short_code"],
        "scheme_name": scheme_info["scheme_name"],
        "holdings": holdings,
    }


def run_etl(excel_path: str) -> dict:
    """Run ETL process on Kotak Mahindra MF Excel file."""
    log.info(f"Loading workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # Parse Scheme sheet for abbreviation mapping
    scheme_mapping = parse_scheme_sheet(wb)
    
    # Extract report date from first data sheet header
    report_date = None
    if len(wb.sheetnames) > 0:
        first_sheet = wb[wb.sheetnames[0]]
        first_rows = list(first_sheet.iter_rows(max_row=3, values_only=True))
        for r in first_rows:
            for cell in r:
                if cell and isinstance(cell, str) and "as on" in cell:
                    # Extract date from string like "Portfolio of Kotak ... as on 31-Jan-2026"
                    parts = cell.split("as on")
                    if len(parts) > 1:
                        date_str = parts[1].strip()
                        report_date = parse_date(date_str)
                        break
            if report_date:
                break
    
    log.info(f"Report date: {report_date}")
    
    # Process each scheme sheet (skip "Common Notes" and "Scheme")
    all_holdings = []
    funds_with_equity = []
    security_counter = Counter()
    skipped = 0
    errors = 0
    
    for sheet_name in wb.sheetnames:
        # Skip utility sheets
        if sheet_name in ["Common Notes", "Scheme"]:
            continue
        
        try:
            ws = wb[sheet_name]
            ws_rows = list(ws.iter_rows(values_only=True))
            
            # Get full scheme name from mapping, or use sheet name
            full_name = scheme_mapping.get(sheet_name, sheet_name)
            
            scheme_info = {
                "scheme_short_code": sheet_name,
                "scheme_name": full_name,
            }
            
            # Extract equity holdings
            result = extract_equity_holdings(ws_rows, scheme_info)
            
            if result is None:
                skipped += 1
                log.debug(f"  {sheet_name}: No equity data")
                continue
            
            num_holdings = len(result["holdings"])
            log.info(f"  {sheet_name}: Extracted {num_holdings} equity holdings")
            
            # Add to funds with equity
            funds_with_equity.append({
                "scheme_short_code": sheet_name,
                "scheme_name": full_name,
                "scheme_code": sheet_name,
                "holdings_count": num_holdings,
            })
            
            # Process holdings
            for h in result["holdings"]:
                h["scheme_short_code"] = sheet_name
                all_holdings.append(h)
                security_counter[h["isin"]] += 1
        
        except Exception as e:
            log.error(f"  {sheet_name}: Error - {e}")
            errors += 1
    
    # Build security master (deduplicated by ISIN)
    security_master = []
    isin_to_info = {}
    
    for h in all_holdings:
        isin = h["isin"]
        if isin not in isin_to_info:
            isin_to_info[isin] = {
                "isin": isin,
                "security_name": h["security_name"],
                "current_industry": h["industry"],
                "current_sector": None,  # Kotak doesn't have separate sector
            }
    
    security_master = list(isin_to_info.values())
    
    # Build output
    output = {
        "metadata": {
            "source_file": os.path.basename(excel_path),
            "report_date": report_date,
            "extraction_date": datetime.now().strftime("%Y-%m-%d"),
            "total_schemes": len(wb.sheetnames) - 2,  # Exclude Common Notes and Scheme
            "schemes_with_equity": len(funds_with_equity),
            "schemes_skipped": skipped,
            "errors": errors,
            "total_unique_securities": len(security_master),
            "total_holdings_records": len(all_holdings),
        },
        "amc_master": {
            "amc_name": "Kotak Mahindra Mutual Fund",
            "short_code": "KOTAKMF",
        },
        "fund_master": funds_with_equity,
        "security_master": security_master,
        "portfolio_holdings": all_holdings,
    }
    
    # Summary
    log.info("=" * 60)
    log.info("ETL COMPLETE")
    log.info(f"  Schemes processed: {len(funds_with_equity)}/{len(wb.sheetnames)-2}")
    log.info(f"  Unique securities: {len(security_master)}")
    log.info(f"  Total holdings:    {len(all_holdings)}")
    log.info(f"  Skipped:           {skipped}")
    log.info(f"  Errors:            {errors}")
    log.info("=" * 60)
    
    return output


def main():
    parser = argparse.ArgumentParser(
        description="Extract equity holdings from Kotak Mahindra MF portfolio Excel"
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")
    
    args = parser.parse_args()
    
    if args.verbose:
        log.setLevel(logging.DEBUG)
    
    if not os.path.exists(args.excel_file):
        log.error(f"File not found: {args.excel_file}")
        sys.exit(1)
    
    # Run ETL
    result = run_etl(args.excel_file)
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        # Auto-generate based on report date
        report_date = result["metadata"]["report_date"]
        if report_date:
            date_suffix = report_date.replace("-", "")[:6]  # YYYYMM
        else:
            date_suffix = datetime.now().strftime("%Y%m")
        
        output_dir = Path("data/processed")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"kotak_equity_holdings_{date_suffix}.json"
    
    # Write JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    
    log.info(f"\n[OK] Output written to: {output_path}")


if __name__ == "__main__":
    main()
