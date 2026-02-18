"""
ETL script for Bajaj Finserv Mutual Fund monthly portfolio Excel files.
Extracts Listed Equity holdings only and outputs JSON matching database schema.

Sheet structure (no Index sheet):
  Row 1:  [sheet_code, scheme_full_name, ...]
  Row 3:  [..., 'Monthly Portfolio Statement as on DD Mon YYYY', ...]
  Row 4:  Headers: [code, Name of Instrument, ISIN, Industry, Quantity, Market Value, % to Net Assets, YTM]
  Row 5:  'Equity & Equity related'  (in col 1)
  Row 6:  '(a) Listed / awaiting listing on Stock Exchanges'  (in col 1)
  Row 7+: Data rows
  End:    'Sub Total' / '(b)' / 'TOTAL' in col 1
"""

import argparse
import json
import logging
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

# Column indices (0-indexed)
COL_NAME = 1          # Name of the Instrument
COL_ISIN = 2          # ISIN
COL_INDUSTRY = 3      # Industry / Rating
COL_QUANTITY = 4      # Quantity
COL_MARKET_VALUE = 5  # Market/Fair Value (Rs. in Lakhs)
COL_PCT_NAV = 6       # % to Net Assets

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("bajaj_etl")


def safe_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        v = float(val)
        return v if v != 0 else None
    except (ValueError, TypeError):
        return None


def safe_int(val: Any) -> int | None:
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def clean_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_date(val: Any) -> str | None:
    """Parse date from text like 'Monthly Portfolio Statement as on 31 Jan 2026'."""
    if val is None:
        return None
    s = str(val).strip()

    # Pattern: "as on DD Mon YYYY"
    m = re.search(r'as\s+on\s+(\d{1,2})\s+(\w+)\s+(\d{4})', s, re.IGNORECASE)
    if m:
        day, mon, year = m.group(1), m.group(2), m.group(3)
        for fmt in ("%d %b %Y", "%d %B %Y"):
            try:
                return datetime.strptime(f"{day} {mon} {year}", fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue

    # Pattern: "DD-Mon-YYYY"
    m = re.search(r'(\d{1,2})-(\w{3})-(\d{4})', s)
    if m:
        try:
            return datetime.strptime(m.group(0), "%d-%b-%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass

    return None


def extract_equity_holdings(ws_rows: list[tuple], scheme_info: dict) -> dict | None:
    """Extract ONLY listed equity holdings from a Bajaj Finserv scheme sheet.

    Equity section structure (all in col 1):
      Row N:   'Equity & Equity related'
      Row N+1: '(a) Listed / awaiting listing on Stock Exchanges'
      Row N+2+: data rows
      Row M:   'Sub Total' or '(b)' or 'TOTAL'  ← stop here

    For debt/hybrid funds, equity section may be empty (Sub Total immediately follows).
    """
    scheme_name = scheme_info.get("scheme_name", "Unknown")

    # Find equity header in col 1
    equity_row = None
    for i, r in enumerate(ws_rows):
        cell1 = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        if not cell1:
            continue
        cl = cell1.lower()
        if 'equity' in cl and 'related' in cl:
            equity_row = i
            break

    if equity_row is None:
        return None

    # Look for '(a) Listed' within 3 rows of equity header (col 1)
    NON_EQUITY = ('debt', 'money market', 'government', 'securitised', 'treps', 'derivatives')
    listed_row = None

    for i in range(equity_row + 1, min(equity_row + 4, len(ws_rows))):
        r = ws_rows[i]
        cell1 = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        if not cell1:
            continue
        cl = cell1.lower()

        if any(sec in cl for sec in NON_EQUITY):
            log.debug(f"  {scheme_name}: No listed equity (found '{cell1}' after equity header)")
            return None

        if '(a)' in cl and ('listed' in cl or 'awaiting' in cl):
            listed_row = i
            break

    if listed_row is None:
        log.debug(f"  {scheme_name}: No '(a) Listed' subsection found near equity header")
        return None

    # Extract holdings until Sub Total / (b) / TOTAL
    holdings = []
    isin_map = {}

    for i in range(listed_row + 1, len(ws_rows)):
        r = ws_rows[i]

        cell1 = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        if cell1:
            cl = cell1.lower().strip()
            if cl in ('sub total', 'total', 'sub-total'):
                break
            if cl.startswith('(b)') or cl.startswith('(c)'):
                break

        name = clean_str(r[COL_NAME]) if len(r) > COL_NAME else None
        isin = clean_str(r[COL_ISIN]) if len(r) > COL_ISIN else None
        industry = clean_str(r[COL_INDUSTRY]) if len(r) > COL_INDUSTRY else None
        quantity = safe_float(r[COL_QUANTITY]) if len(r) > COL_QUANTITY else None
        market_value = safe_float(r[COL_MARKET_VALUE]) if len(r) > COL_MARKET_VALUE else None
        pct_nav = safe_float(r[COL_PCT_NAV]) if len(r) > COL_PCT_NAV else None

        if not isin or not name:
            continue

        if len(isin) != 12 or not isin.startswith("INE"):
            log.debug(f"  Skipping invalid ISIN: {isin}")
            continue

        if isin in isin_map:
            idx = isin_map[isin]
            existing = holdings[idx]
            if quantity and existing.get("quantity"):
                existing["quantity"] += quantity
            if market_value and existing.get("market_value_lakhs"):
                existing["market_value_lakhs"] += market_value
            continue

        holding = {
            "security_name": name,
            "isin": isin,
            "industry": industry,
            "quantity": safe_int(quantity) if quantity else None,
            "market_value_lakhs": round(market_value, 4) if market_value else None,
            "pct_to_nav": round(pct_nav, 4) if pct_nav else None,
        }

        isin_map[isin] = len(holdings)
        holdings.append(holding)

    if not holdings:
        return None

    return {
        "scheme_name": scheme_name,
        "scheme_code": scheme_info.get("sheet_code", ""),
        "holdings": holdings,
    }


def run_etl(excel_path: str, date_override: str | None = None) -> dict:
    """Run the ETL process on a Bajaj Finserv portfolio Excel file."""
    log.info(f"Processing: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    all_funds = []
    all_securities = {}
    all_holdings = []
    report_date = date_override
    skipped = 0

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            ws_rows = list(ws.iter_rows(values_only=True))

            # Extract scheme name from B1 (row 0, col 1)
            scheme_name = None
            if len(ws_rows) > 0 and len(ws_rows[0]) > 1:
                scheme_name = clean_str(ws_rows[0][1])
            if not scheme_name:
                scheme_name = sheet_name

            # Extract report date from Row 3 col 1 (first sheet only)
            if report_date is None and len(ws_rows) > 2:
                for col_idx in range(min(4, len(ws_rows[2]))):
                    parsed = parse_date(ws_rows[2][col_idx])
                    if parsed:
                        report_date = parsed
                        log.info(f"Report date: {report_date}")
                        break

            result = extract_equity_holdings(ws_rows, {
                "scheme_name": scheme_name,
                "sheet_code": sheet_name,
            })

            if not result:
                skipped += 1
                continue

            holdings = result["holdings"]

            fund_record = {
                "scheme_name": result["scheme_name"],
                "scheme_code": sheet_name,
                "scheme_short_code": sheet_name,
            }
            all_funds.append(fund_record)

            for h in holdings:
                isin = h["isin"]
                if isin not in all_securities:
                    all_securities[isin] = {
                        "isin": isin,
                        "security_name": h["security_name"],
                        "current_industry": h.get("industry"),
                        "current_sector": None,
                    }

                all_holdings.append({
                    "scheme_short_code": sheet_name,
                    "isin": isin,
                    "report_date": report_date,
                    "quantity": h.get("quantity"),
                    "market_value_lakhs": h.get("market_value_lakhs"),
                    "pct_to_aum": h.get("pct_to_nav"),
                    "industry": h.get("industry"),
                })

            log.info(f"  ✓ {sheet_name}: {len(holdings)} holdings")

        except Exception as e:
            log.warning(f"  ✗ {sheet_name}: {e}")
            skipped += 1

    if not report_date:
        log.warning("Could not extract report date, using current date")
        report_date = datetime.now().strftime("%Y-%m-%d")

    log.info(f"  Schemes: {len(all_funds)}, Skipped: {skipped}")
    log.info(f"  Securities: {len(all_securities)}")
    log.info(f"  Total holdings: {len(all_holdings)}")

    return {
        "metadata": {
            "source_file": os.path.basename(excel_path),
            "amc": "Bajaj Finserv Mutual Fund",
            "report_date": report_date,
            "total_sheets": len(wb.sheetnames),
            "schemes_with_equity": len(all_funds),
            "total_unique_securities": len(all_securities),
            "total_holdings_records": len(all_holdings),
        },
        "amc_master": {
            "amc_name": "Bajaj Finserv Mutual Fund",
            "short_code": "BAJAJFINSERV",
        },
        "fund_master": all_funds,
        "security_master": list(all_securities.values()),
        "portfolio_holdings": all_holdings,
    }


def main():
    parser = argparse.ArgumentParser(
        description="ETL for Bajaj Finserv Mutual Fund portfolio Excel files"
    )
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    parser.add_argument("--date", "-d", help="Override report date (YYYY-MM-DD)")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    data = run_etl(args.excel_file, date_override=args.date)

    if args.output:
        out_path = args.output
    else:
        stem = Path(args.excel_file).stem.lower()
        out_path = f"data/processed/bajaj/bajaj_equity_holdings_{stem}.json"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    log.info(f"Written to: {out_path}")


if __name__ == "__main__":
    main()
